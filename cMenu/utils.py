from typing import (Dict, List, Any, Tuple, NamedTuple )
from collections import (namedtuple, )
from collections.abc import (Callable, )

from django.db import models
from django.db.models import (QuerySet, Model)
from django.db.models import Aggregate, CharField
from django.db.backends.utils import CursorWrapper

from PySide6.QtCore import (QCoreApplication,
    Qt, Slot, QObject,
    QRect,
    QStringListModel, QAbstractTableModel, QAbstractListModel, QModelIndex,
    )
from PySide6.QtWidgets import (QApplication, QWidget,
    QMessageBox, 
    QVBoxLayout, QScrollArea, QFrame, QGridLayout,
    QLineEdit, QCompleter, 
    QComboBox, QPushButton,
    )
# from PySide6.QtGui import (
#    )

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, fills, colors
from openpyxl.utils.datetime import from_excel, WINDOWS_EPOCH


ExcelWorkbook_fileext = ".XLSX"

def pleaseWriteMe(parent, addlmessage):
    msg = QMessageBox(parent)
    msg.setWindowTitle('Please Write Me')
    msg.setIcon(QMessageBox.Icon.Warning)
    msg.setStandardButtons(QMessageBox.StandardButton.Ok)
    msg.setText(f'Calvin needs to get up off his butt and write some code\n{addlmessage}')
    msg.open()


class cDataList(QLineEdit):
    """
    A LineEdit box that acts like an HTML DataList
    Choice matches are choices which contain the input string, case-insensitive
    
    caller should connect the editingFinished signal to a slot which is aware of the cDataList
        and is in scope to call cDataList.selectedItem
    ex: self.testdatalist.editingFinished.connect(self.showHBLChosen)

    Args:
        choices:Dict[Any, str], {key: 'value to display/lookup'}
        initval:str = '', (not currently implemented)
        parent:QWidget=None

    def selectedItem(self):
        returns the following dictionary: 
        return {'keys': [key for key, t in self.choices.items() if t==self.text()], 'text': self.text()}
        (all keys matching the text input)
    """
    def __init__(self, choices:Dict[Any, str], initval:str = '', parent:QWidget=None):
        super().__init__(initval, parent)

        self.choices = choices
        
        self.setClearButtonEnabled(True)
        
        choices_to_present = list(choices.values())
        qCompleterObj = QCompleter(QStringListModel(choices_to_present, self), self)
        qCompleterObj.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        qCompleterObj.setFilterMode(Qt.MatchFlag.MatchContains)
        qCompleterObj.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        
        self.setCompleter(qCompleterObj)
    # __ init__
    
    def selectedItem(self):
        """
        def selectedItem(self):
            returns the following dictionary: 
            return {'keys': [key for key, t in self.choices.items() if t==self.text()], 'text': self.text()}
            (all keys matching the text input)
        """
        return {'keys': [key for key, t in self.choices.items() if t==self.text()], 'text': self.text()}
    # selectedItem
    
    def addChoices(self, choices:Dict[Any, str]):
        self.choices.update(choices)
        
        choices_to_present = list(self.choices.values())
        newmodel = QStringListModel(choices_to_present, self)
        self.completer().setModel(newmodel)
    # addChoices




from PySide6.QtCore import QAbstractTableModel, Qt

class cDictModel(QAbstractTableModel):
    def __init__(self, data, parent=None):
        """
        Initialize the model with a dictionary.
        
        Args:
            data (dict): Dictionary to model.
        """
        super().__init__(parent)
        self._data = data
        self._keys = list(data.keys())

    def rowCount(self, parent=None):
        """Return the number of rows in the model."""
        return len(self._data)

    def columnCount(self, parent=None):
        """Return the number of columns in the model (Key and Value)."""
        return 2  # One for the key and one for the value

    def data(self, index, role=Qt.DisplayRole):
        """Return the data for a given cell."""
        if not index.isValid() or role != Qt.DisplayRole:
            return None

        row = index.row()
        col = index.column()

        key = self._keys[row]
        if col == 0:  # Key column
            return key
        elif col == 1:  # Value column
            return self._data[key]
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        """Return the header labels for rows or columns."""
        if role != Qt.DisplayRole:
            return None

        if orientation == Qt.Horizontal:
            return ["Key", "Value"][section]  # Column headers
        elif orientation == Qt.Vertical:
            return str(section + 1)  # Row numbers
        return None

    def setData(self, index, value, role=Qt.EditRole):
        """Set the data for a given cell."""
        if not index.isValid() or role != Qt.EditRole:
            return False

        row = index.row()
        col = index.column()
        key = self._keys[row]

        if col == 1:  # Only allow editing the value column
            self._data[key] = value
            self.dataChanged.emit(index, index, [role])
            return True

        return False

    def flags(self, index):
        """Set flags for each cell."""
        if not index.isValid():
            return Qt.NoItemFlags

        flags = Qt.ItemIsSelectable | Qt.ItemIsEnabled
        if index.column() == 1:  # Only value column is editable
            flags |= Qt.ItemIsEditable

        return flags

class cComboBoxFromDict(QComboBox):
    """
    Generates QComboBox from dictionary

    Args:
        dict (Dict): The input dictionary. The values will be the data returned by currentData, and
            the keys will be the values shown in the QComboBox
        parent (QWidget) default None
    
    """
    _combolist:List = []
    
    def __init__(self, dict:Dict, parent:QWidget = None):
        super().__init__(parent)
        
        # don't do completers - assume underlying QComboBox is non-editable
        # choices_to_present = list(dict)
        # qCompleterObj = QCompleter(QStringListModel(choices_to_present, self), self)
        # qCompleterObj.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        # qCompleterObj.setFilterMode(Qt.MatchFlag.MatchContains)
        # qCompleterObj.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        # self.setCompleter(qCompleterObj)

        for key, val in dict.items():
            self.addItem(key,val)
            self._combolist.append({key:val})

#########################################        
#########################################        

# cQRecordsetView - Scrollable layout of records
class cQRecordsetView(QWidget):
    _newdgt_fn:Callable[[], QWidget] = None
    _btnAdd:QPushButton = None
    def __init__(self, newwidget_fn:Callable[[], QWidget] = None, parent=None):
        """
        Widget which displays a set of records

        Args:
            newwidget_fn (Callable, optional): . Defaults to None. newwidget_fn() should return a new record 
                in a widget suitable for adding to this layout
            parent (_type_, optional): _description_. Defaults to None.
        """
        super().__init__(parent)
        self._newdgt_fn = newwidget_fn
        self.init_ui()

    def init_ui(self):
        self.mainLayout = QVBoxLayout(self)

        # set up scroll area
        self.scrollarea = QScrollArea(self)
        self.scrollarea.setWidgetResizable(True)
        self.scrollarea.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.scrollarea.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.mainLayout.addWidget(self.scrollarea)
    
        # Container widget for the layout
        self.scrollwidget = QWidget()
        # layout for the scrollwidget
        self.scrolllayout = QVBoxLayout(self.scrollwidget)
        # # Set container as the scroll area widget
        self.scrollarea.setWidget(self.scrollwidget)

        if self._newdgt_fn:
            self._btnAdd = QPushButton(self.scrollwidget)
            self._btnAdd.setObjectName('AddBtnQRS')
            self._btnAdd.setText('\nAdd\n')
            self._btnAdd.clicked.connect(self.addBtnClicked)
            self.scrolllayout.addWidget(self._btnAdd)

        self.init_recSet()
    # init_ui

    def setAddText(self, addText:str = '\nAdd\n'):
        ...
    # setAddText

    def addWidget(self, wdgt:QWidget):
        insAt = self.scrolllayout.count()-1 if self._newdgt_fn else -1
        self.scrolllayout.insertWidget(insAt, wdgt)
        line = QFrame(self)
        myWidth = self.geometry().width()
        line.setGeometry(QRect(0, 0, myWidth, 3))
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        self.scrolllayout.insertWidget(insAt+1, line)
    # addWidget

    # addLayout needed?

    def init_recSet(self):
        # remove all widgets from scrolllayout
        # for wdgt in self.scrolllayout:
        idx = 0
        child = self.scrolllayout.itemAt(idx)
        while child != None:
            if child.widget() == self._btnAdd:   # don't removew the Add Button!
                idx += 1
            else:
                widg = self.scrolllayout.takeAt(idx)
                widg.widget().deleteLater() # del the widget
            # endif child == self._btnAdd
            child = self.scrolllayout.itemAt(idx)
    # init_recSet
    
    @Slot()
    def addBtnClicked(self):
        if callable(self._newdgt_fn):
            self.addWidget(self._newdgt_fn())
    # addBtnClicked

#########################################        
#########################################        

class QDjangoTableModel(QAbstractTableModel):
    def __init__(self, tbl:models.Model, flds:List, foreign_keys:List = [], special_processing_flds:Dict[str,Tuple[Callable, Callable]] = {}, filter:Dict[str,Any] = {}, parent = None):
        super().__init__(parent)
        self.headers = flds
        self.queryset = list(tbl.objects.filter(**filter).only(*flds))
        self.foreign_keys = foreign_keys
        self.special_processing = special_processing_flds

    
    def rowCount(self, parent = QModelIndex()):
        return len(self.queryset)
    
    def columnCount(self, parent = QModelIndex()):
        return len(self.headers)
    
    def data(self, index, role = Qt.DisplayRole):
        if not index.isValid():
            return None
        if role == Qt.ItemDataRole.DisplayRole:
            rec = self.queryset[index.row()]
            fldName = self.headers[index.column()]
            value = getattr(rec, fldName, '')
            if fldName in self.foreign_keys:
                value = str(value) if value else ''
            elif fldName in self.special_processing:
                formatter, _ = self.special_processing[fldName]
                if callable(formatter):
                    return formatter(value)
            # endif special field values

            return value
        # endif role
        return None

    def setData(self, index, value, role=Qt.EditRole):
        if not index.isValid():
            return False
        if role == Qt.EditRole:
            rec = self.queryset[index.row()]
            fldName = self.headers[index.column()]

            # Apply special processing for editing if defined
            if fldName in self.special_processing:
                _, editor = self.special_processing[fldName]
                if callable(editor):
                    value = editor(value)

            # Set the field value
            setattr(rec, fldName, value)
            rec.save()
            return True
        return False

 
    def headerData(self, section, orientation, role = Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                return self.headers[section]
            elif orientation == Qt.Orientation.Vertical:
                return str(section+1)
            #endif orientation
        # endif role
        return None


#########################################        
#########################################        

class NOPETHATSNOTIT(QAbstractListModel):
    def __init__(self, data:Dict[str,Any], parent:QObject=None):
        """
        Initialize the model with a list of dictionaries.
        
        Args:
            data (list of dict): List of dictionaries to model.
        """
        super().__init__(parent)
        self._data = data

    def rowCount(self, parent=None) -> int:
        """Return the number of rows in the model."""
        return len(self._data)

    # def columnCount(self, parent=None) -> int:
    #     """Return the number of columns in the model."""
    #     return len(self._headers)

    def data(self, index:QModelIndex, role:Qt.ItemDataRole=Qt.DisplayRole) -> Any|None:
        """Return the data for a given cell."""
        if not index.isValid() or role != Qt.DisplayRole:
            return None

        row = index.row()
        col = index.column()
        key = self._headers[col]
        return self._data[row].get(key, None)

    def headerData(self, section:int, orientation:Qt.Orientation, role:Qt.ItemDataRole=Qt.DisplayRole) -> str|int|None:
        """Return the header labels for rows or columns."""
        if role != Qt.DisplayRole:
            return None

        if orientation == Qt.Horizontal:
            return self._headers[section]  # Column headers (dictionary keys)
        elif orientation == Qt.Vertical:
            return str(section + 1)  # Row numbers
        return None

    def setData(self, index:QModelIndex, value:Any, role:Qt.ItemDataRole=Qt.EditRole) -> bool:
        """Set the data for a given cell."""
        if not index.isValid() or role != Qt.EditRole:
            return False

        row = index.row()
        col = index.column()
        key = self._headers[col]
        self._data[row][key] = value
        self.dataChanged.emit(index, index, [role])
        return True

    def flags(self, index:QModelIndex) -> Qt.ItemFlag:
        """Set flags for each cell."""
        if not index.isValid():
            return Qt.NoItemFlags

        return Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable


#####################################################
#####################################################
#####################################################


def dictfetchall(cursor:CursorWrapper) -> List[Dict[str,Any]]:
    """
    Return all rows from a cursor as a list of dictionaries.
    Assume the column names are unique.
    """
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

def namedtuplefetchall(cursor:CursorWrapper, ResultName:str = 'Result') -> List[NamedTuple]:
    "Return all rows from a cursor as a namedtuple"
    desc = cursor.description
    nt_result = namedtuple(ResultName, [col[0] for col in desc])
    return [nt_result(*row) for row in cursor.fetchall()]

def modelobj_to_dict(modelobj:Model) -> Dict[str, Any]:
    # opts = modelobj._meta
    # data = {}
    # for f in chain(opts.concrete_fields, opts.private_fields):
    #     data[f.name] = f.value_from_object(modelobj)
    # for f in opts.many_to_many:
    #     data[f.name] = [i.id for i in f.value_from_object(modelobj)]

    data = {key:val for key, val in modelobj.__dict__.items() if key not in ['_state']}
    return data

def Excelfile_fromqs(qset:QuerySet|List[Dict[str, Any]], flName:str|None = None, 
                     freezecols:int = 0, returnFileName: bool = False) -> Workbook|str:
    """
    qset: a queryset or list of dictionaries
    flName: the name of the file to be built (WITHOUT extension!).  It's stored on the server.  If it's to be dl'd, the caller does that
    freezecols = 0: the number of columns to freeze to the left
    The top row contains the field names, is always frozen, is bold and is shaded grey

    used to Return the name of the Workbook file (with extension).  Once I start building in errorchecking and exceptions, other returns may be possible
    Returns the Workbook file (with extension)
    """

    # far easier to process a list of dictionaries, so...
    if isinstance(qset,QuerySet):
        qlist = qset.values()
    elif isinstance(qset,list):
        qlist = qset
    else:
        return None
    if qlist:
        if not isinstance(qlist[0],dict):
            # review this later ...
            try:
                qlist = [n.__dict__ for n in qlist]
            except:
                qlist = []

    # create empty workbook with an empty worksheet
    wb = Workbook()
    ws = wb.active

    # header row is names of columns
    if qlist:
        fields = list(qlist[0])
        ws.append(fields)

        # append each row
        for row in qlist:
            ws.append(list(row.values()))

        # make header row bold, shade it grey, freeze it
        # ws.show_gridlines = True  #Nope - this is a R/O attribute
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type=fills.FILL_SOLID,
                            start_color=colors.Color("00808080"),
                            end_color=colors.Color("00808080")
                            )
        #TODO: convert row1 and cols:freezecols to an address (A=0, B=1, C=2 etc) for line below
        ws.freeze_panes ='A2'
        #TODO: if freezecols passed, freeze them, too


    # save the workbook
    if flName:
        wb.save(flName+ExcelWorkbook_fileext)
    
    if returnFileName:
        # close the workbook
        wb.close()
        # and return file Name to the caller
        return flName+ExcelWorkbook_fileext
    else:
        # return the workbook itself
        return wb
    #endif returnFileName

class GroupConcat(Aggregate):
    function = 'GROUP_CONCAT'
    allow_distinct = True
    template = '%(function)s(%(distinct)s%(expressions)s%(ordering)s%(separator)s)'

    def __init__(self, expression, distinct=False, ordering=None, separator=None, **extra):
        super().__init__(
            expression,
            distinct='DISTINCT ' if distinct else '',
            ordering=' ORDER BY %s' % ordering if ordering is not None else '',
            separator=' SEPARATOR "%s"' % separator if separator is not None else '',
            output_field=CharField(),
            **extra
        )


class UpldSprdsheet():
    TargetModel = None
    SprdsheetDateEpoch = WINDOWS_EPOCH

    def SprdsheetFldDescriptor_creator(self, ModelFldName, AllowedTypes):
        """
        ModelFldName (str): the name of the field in the TargetModel
        AllowedTypes: list of tuples (type, cleanproc).  empty list if any string allowed
        """
        return  {
            # 'SprdsheetName': None,    # nope, this will be the index of SprdsheetFlds
            'ModelFldName': ModelFldName,
            'AllowedTypes': AllowedTypes,     
        }
    
    SprdsheetFlds = {}  # key will be the SprdsheetName, value is a SprdsheetFldDescriptor

    def cleanupfld(self, fld, val):
        usefld = False
        cleanval = None
        
        if fld not in self.SprdsheetFlds:
            # just feed the value back
            usefld = True
            cleanval = val
        elif not self.SprdsheetFlds[fld]['AllowedTypes']:
            usefld = (val is not None)
            if usefld: cleanval = str(val)
        else:
            for type, cleanproc in self.SprdsheetFlds[fld]['AllowedTypes']:
                if isinstance(val, type):
                    usefld = True
                    cleanval = cleanproc(val)
                    break
                #endif instance(val, type)
            #endfor type, cleanproc
        #endif fld not in self.SprdsheetFlds

    def process_spreadsheet(self, SprsheetName):
        pass

