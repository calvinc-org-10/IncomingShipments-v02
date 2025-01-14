from typing import Any, Dict, List

import re
from openpyxl import load_workbook
from datetime import date
from django.http import HttpResponse
from incShip.models import HBL, ShippingForms, Containers, Invoices, references, reference_ties
from incShip.models import Companies, FreightTypes

from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
    QFont, QFontDatabase, QGradient, QIcon,
    QImage, QKeySequence, QLinearGradient, QPainter,
    QPalette, QPixmap, QRadialGradient, QTransform)
from PySide6.QtWidgets import (QApplication, QWidget, 
    QGridLayout, QSpacerItem,
    QListView, QListWidget, QColumnView, QListWidgetItem, QTableView, QHeaderView,
    QComboBox, QDateEdit, QFrame,
    QLabel, QLineEdit, 
    QPushButton,
    QSizePolicy, QTabWidget, QTextEdit,
    )

from forms import std_popdialogsize

class LoadHBL(QWidget):
    _linkedTables = ['ShippingForms', 'PO', 'Invoices', 'Containers', 'reference_ties']
    lblFormName = QLabel()

    lblStatus = QLabel()

    def __init__(self, parent:QWidget = None):
        super().__init__(parent)

        if not self.objectName():
            self.setObjectName(u"Form")
        self.resize(2*std_popdialogsize)

        wdgt = self.lblFormName
        wdgt.setParent(self)
        wdgt.setObjectName(u"lblFormName")
        wdgt.setGeometry(QRect(200, 10, 319, 74))
        #TODO: universal form title font
        font1 = QFont()
        font1.setFamilies([u"Century Gothic"])
        font1.setPointSize(24)
        wdgt.setFont(font1)
        wdgt.setFrameShape(QFrame.Shape.Panel)
        wdgt.setFrameShadow(QFrame.Shadow.Raised)
        wdgt.setAlignment(Qt.AlignmentFlag.AlignCenter)
        wdgt.setWordWrap(True)
        
        wdgt = self.lblStatus
        wdgt.setParent(self)
        wdgt.setWordWrap(True)
        wdgt.setGeometry(10, 30, 600, 400)

        self.retranslateUi()

        self.show()
        self.load_recs()
    # __init__

    def retranslateUi(self):
        self.setWindowTitle(QCoreApplication.translate("Form", "Load HBL from SprSht", None))
        self.lblFormName.setText(QCoreApplication.translate("Form", "Load HBL from SprSht", None))
   # retranslateUi

    def load_recs(self):
        fName = f'W:\\Logistics\\Invoices\\Invoices.xlsx'
        shtName = 'HBL heads-ups'
        wb = load_workbook(filename=fName, read_only=True)
        ws = wb[shtName]
        
        HBL_spsht = ws
        
        process_flag = False
        import_spsht_prfx='SpSht.Import.2024-12-28'
        sentinel_col, sentinel_flag = ('LstDigChk', 'RESTART')
        
        colmnNames = ws[1]
        colmnMap = { }    # start with columns that MUST be in the SpSht
        SSName_TableName_map = {
            sentinel_col: None ,         # used only to find the sentinel
            'HBL': 'HBLNumber', 
            'sh fm': None,              # mod/create ShippingForm
            'container num': None,      # mod/create a Container
            'mode': 'FreightType', 
            'origin': 'Origin', 
            'notes': 'notes',           # in the refs table
            'ETA ORD or MKE': 'ETA', 
            'Deliv appt': 'DelivAppt',  # in Containers table
            'LFD': 'LFD',               # in HBL AND CXontainers
            'orig invc': 'InvoiceNumber',  # mod/create Invoices
            'orig SmOff': 'id_SmOffFormNum',  # mod/create Invoices
            'Company': 'Company',
            }
        # map column names to column numbers
        for col in colmnNames:
            if col.value in SSName_TableName_map:
                colkey = col.value # SSName_TableName_map[col.value] - not used - I do the spSht -> Table mappings manually here
                # has this col.value already been mapped?
                if (colkey in colmnMap and colmnMap[colkey] is not None):
                    # yes, that's a problem
                    statecode = 'fatalerr'
                    statetext = f'SAP Spreadsheet has bad header row - More than one column named {col.value}.  See Calvin to fix this.'
                    result = 'FAIL - bad spreadsheet'
                    self.lblStatus.setText(f'{statecode}\n{statetext}\n{result}')
                    wb.close()
                    return
                else:
                    colmnMap[colkey] = col.column - 1
                # endif previously mapped
            #endif col.value in SAP_SSName_TableName_map
        #endfor col in SAPcolmnNames
        
        numrows = ws.max_row
        numrow98 = .98 * numrows
        numrow90 = .90 * numrows
        SprshtRowNum = 1
        announceInterval = 100
        
        Company_dflt =  Companies.objects.get(pk=1)    # not always true, but I'll fix individual cases manually
        
        for row in ws.iter_rows(min_row=SprshtRowNum+1, values_only=True):
            SprshtRowNum += 1
            
            if not process_flag:
                if row[colmnMap[sentinel_col]] == sentinel_flag:
                    process_flag = True
                continue
            # endif process_flag

            
            HBLNum = row[colmnMap['HBL']]
            ShFm = row[colmnMap['sh fm']]
            ContNum = row[colmnMap['container num']]
            InvNum = row[colmnMap['orig invc']]
            print(SprshtRowNum)
            Company = Companies.objects.get(CompanyName__istartswith=row[colmnMap['Company']]) if row[colmnMap['Company']] else Company_dflt
            if InvNum:
                if re.match(r'bill.*', InvNum): InvNum = None

            ETA = row[colmnMap['ETA ORD or MKE']]
            LFD = row[colmnMap['LFD']]
            #create records
            HBLrec = None
            if HBLNum:
                HBLrec, creatFlag =  HBL.objects.get_or_create(HBLNumber = HBLNum)
                HBLrec.Company = Company
                if ETA:
                    HBLrec.ETA = ETA
                if LFD:
                    HBLrec.LFD = LFD
                HBLrec.save()
            ShFmrec = None
            if ShFm:
                ShFmrec, creatFlag = ShippingForms.objects.get_or_create(id_SmOffFormNum = ShFm)
            Contrec = None
            if ContNum:
                Contrec, creatFlag = Containers.objects.get_or_create(ContainerNumber = ContNum)
                Contrec.DelivAppt = row[colmnMap['Deliv appt']]
                if LFD:
                    Contrec.LFD = LFD
                Contrec.save()
            Invrec = None
            if InvNum and HBLNum:
                FType = FreightTypes.objects.get(pk=1)
                v = row[colmnMap['mode']]
                if v:
                    if re.search(r'air', v,re.IGNORECASE): FType = FreightTypes.objects.filter(FreightType__icontains='air').first()
                    if re.search(r'ocean', v, re.IGNORECASE): FType = FreightTypes.objects.filter(FreightType__icontains='ocean').first()
                Invrec, creatFlag = Invoices.objects.get_or_create(InvoiceNumber=InvNum,
                    defaults= {
                        'HBL':HBLrec,
                        'SmOffStatus': Invoices.SmOffStatusCodes.PENDING,
                        'InvoiceDate': date.today(),
                        'InvoiceAmount': 0,
                    }
                )
                Invrec.verifiedForFii = True
                Invrec.inv_downloaded = True
                if row[colmnMap['orig SmOff']] and not Invrec.id_SmOffFormNum:
                    Invrec.id_SmOffFormNum = row[colmnMap['orig SmOff']]
                Invrec.save()
            
            if HBLrec and ShFmrec: 
                HBLrec.ShippingForms.add(ShFmrec)
            if HBLrec and Contrec:
                Contrec.HBL = HBLrec
                Contrec.save()
            noterec = None
            if row[colmnMap['notes']]:
                noterec, creatFlag = references.objects.get_or_create(
                    emaildatefrom_or_filelocation = f'{import_spsht_prfx}.{SprshtRowNum}',
                    defaults={
                        'notes': row[colmnMap['notes']]
                    }
                )
                if HBLrec:
                    reference_ties.objects.get_or_create(email=noterec, 
                                table_ref=reference_ties.refTblChoices.HBL,
                                record_ref=HBLrec.pk)
                if ShFmrec:
                    reference_ties.objects.get_or_create(email=noterec, 
                                table_ref=reference_ties.refTblChoices.ShippingForms,
                                record_ref=ShFmrec.pk)
                if Contrec:
                    reference_ties.objects.get_or_create(email=noterec, 
                                table_ref=reference_ties.refTblChoices.Containers,
                                record_ref=Contrec.pk)
                if Invrec:
                    reference_ties.objects.get_or_create(email=noterec, 
                                table_ref=reference_ties.refTblChoices.Invoices,
                                record_ref=Invrec.pk)
            self.lblStatus.setText(f'{SprshtRowNum}, {row}, {HBLNum}, {ShFm}, {ContNum}, {InvNum} processed')

        self.lblStatus.setText('Data pulled from spreadsheet')
    # load_recs

